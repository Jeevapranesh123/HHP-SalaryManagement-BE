from app.database import get_mongo, AsyncIOMotorClient
import pandas as pd
from copy import copy
import io
import openpyxl
from openpyxl.styles import PatternFill
import openpyxl.utils
from openpyxl.styles import Font
from app.api.crud.admin import AdminCrud

from fastapi import HTTPException, Depends, APIRouter

from app.schemas.admin import (
    Rules,
    Guidelines,
    ReportType,
    BankSalaryBatchCreateRequest,
    BankSalaryBatchInDB,
    BankSalaryBatch,
)

import pprint

from app.core import pipelines

import datetime

from app.api.utils import *

from app.api.crud import employees as employee_crud


class AdminController:
    def __init__(self, payload, mongo_client: AsyncIOMotorClient):
        self.payload = payload
        self.mongo_client = mongo_client
        self.employee_id = payload["employee_id"]
        self.primary_role = payload["primary_role"]

    async def post_rules(self, rules: Rules):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        rules = rules.model_dump()
        return await crud_obj.post_rules(rules)

    async def get_rules(self):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        return await crud_obj.get_rules()

    async def get_rules_meta(self):
        meta = {
            "rules": {
                "data": {
                    "total_leave": {
                        "type": "string",
                        "required": True,
                    },
                    "medical_leave": {
                        "type": "string",
                        "required": True,
                    },
                    "permission": {
                        "type": "string",
                        "required": True,
                    },
                    "loan": {
                        "type": "string",
                        "required": True,
                    },
                    "increment": {
                        "type": "string",
                        "required": True,
                    },
                },
                "actions": [
                    {
                        "label": "Update",
                        "type": "button",
                        "variant": "default",
                        "action": {"url": "/admin/rules", "method": "PUT"},
                    }
                ],
            }
        }

        return meta

    async def get_guidelines(self):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        return await crud_obj.get_guidelines()

    async def post_guidelines(self, guidelines: Guidelines):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        guidelines = guidelines.model_dump()

        processed_guidelines = {"guidelines": []}
        for guideline in guidelines["guidelines"]:
            processed_guidelines["guidelines"].append(
                {"id": guideline["id"], "content": guideline["content"].strip().title()}
            )

        return await crud_obj.post_guidelines(processed_guidelines)

    async def get_roles(self):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        roles = await crud_obj.get_roles()
        return roles

    async def get_employee_role(self, employee_id):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        user = await crud_obj.get_employee_role(employee_id)
        if user.get("primary_role"):
            role = user["primary_role"]["role"]

        else:
            role = None

        return role

    async def get_report_meta(self, type=None):
        if not type:
            data = [
                {"label": "Salary", "value": "salary"},
                {"label": "Increment", "value": "increment"},
                {"label": "Bonus", "value": "bonus"},
                {"label": "Allowance", "value": "allowance"},
                {
                    "label": "Attendance Special Allowance",
                    "value": "attendance_special_allowance",
                },
                {
                    "label": "Other Special Allowance",
                    "value": "other_special_allowance",
                },
                {"label": "Leave", "value": "leave"},
                {"label": "All", "value": "all"},
            ]

        elif type != None:
            if type == "salary":
                data = {
                    "salary": {
                        "data": {
                            "employee_id": {"type": "string", "required": True},
                            "month": {
                                "type": "month",
                                "format": "YYYY-MM-DD",
                                "required": True,
                            },
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/salary",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "increment":
                data = {
                    "increment": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/increment",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "bonus":
                data = {
                    "bonus": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/bonus",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "allowance":
                data = {
                    "allowance": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/allowance",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "attendance_special_allowance":
                data = {
                    "attendance_special_allowance": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/attendance_special_allowance",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "other_special_allowance":
                data = {
                    "other_special_allowance": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/other_special_allowance",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "leave":
                data = {
                    "leave": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/leave",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

            elif type == "all":
                data = {
                    "all": {
                        "data": {
                            "start_year": {"type": "year", "required": True},
                            "end_year": {"type": "year", "required": True},
                        },
                        "actions": [
                            {
                                "label": "Submit",
                                "type": "button",
                                "variant": "default",
                                "action": {
                                    "url": "/admin/report/download/all",
                                    "method": "GET",
                                    "isFileRequest": True,
                                },
                            }
                        ],
                    }
                }

        return data

    async def create_bank_salary_batch(self, create_bank_salary_batch):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        md = await employee_crud.get_employee(
            self.payload["employee_id"], self.mongo_client
        )

        if not md:
            raise HTTPException(status_code=404, detail="Requester not found")

        md_branch = md["branch"]

        bank_salary_batch_in_db = BankSalaryBatchInDB(
            **create_bank_salary_batch.dict(), branch=md_branch
        )

        bank_salary_batch_in_db.employee_ids = [
            employee_id.strip() for employee_id in bank_salary_batch_in_db.employee_ids
        ]

        bank_salary_batch_in_db = bank_salary_batch_in_db.model_dump()

        res = await crud_obj.create_bank_salary_batch(bank_salary_batch_in_db)
        res["batch_id"] = res.pop("id")
        return res

    async def get_bank_salary_batch(self, batch_id):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        res = await crud_obj.get_bank_salary_batch(batch_id)
        res["batch_id"] = res.pop("id")

        return res

    async def get_bank_salary_batch_list(self):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        md = await employee_crud.get_employee(
            self.payload["employee_id"], self.mongo_client
        )

        if not md:
            raise HTTPException(status_code=404, detail="Requester not found")

        md_branch = md["branch"]
        res = await crud_obj.get_bank_salary_batch_list(md_branch)
        return res

    async def delete_bank_salary_batch(self, batch_id):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        res = await crud_obj.delete_bank_salary_batch(batch_id)
        return res

    async def update_bank_salary_batch(self, batch_id, update_bank_salary_batch):
        crud_obj = AdminCrud(self.payload, self.mongo_client)

        batch = await crud_obj.get_bank_salary_batch(batch_id)

        if not batch:
            raise HTTPException(status_code=404, detail="Bank salary batch not found")

        # bank_salary_batch_in_db = BankSalaryBatchInDB(**update_bank_salary_batch.dict())

        # bank_salary_batch_in_db.employee_ids = [
        #     employee_id.strip() for employee_id in bank_salary_batch_in_db.employee_ids
        # ]

        # bank_salary_batch_in_db = bank_salary_batch_in_db.model_dump()

        res = await crud_obj.update_bank_salary_batch(
            batch_id, update_bank_salary_batch.model_dump()
        )
        res["batch_id"] = res.pop("id")
        return res

    async def get_bank_salary_batch_list_all(self):
        crud_obj = AdminCrud(self.payload, self.mongo_client)
        md = await employee_crud.get_employee(
            self.payload["employee_id"], self.mongo_client
        )

        if not md:
            raise HTTPException(status_code=404, detail="Requester not found")

        res = await crud_obj.get_bank_salary_batch_list_all(md["branch"])
        return res

    async def copy_cell(self, source_cell, target_cell):
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    async def download_report(self, report_type, query_params):
        if report_type == "bank_salary":
            return await self.download_bank_salary_report(query_params)

        meta = await self.get_report_meta(report_type)

        query_params_check = all(
            key in query_params for key in meta[report_type]["data"].keys()
        )
        print(query_params_check)
        if not query_params_check:
            raise HTTPException(status_code=400, detail="Invalid query parameters")

        if report_type == "salary":
            data = await self.build_salary_report(query_params)
        elif report_type == "increment":
            data = await self.build_increment_report(query_params)
        elif report_type == "bonus":
            data = await self.build_bonus_report(query_params)
        elif report_type == "allowance":
            data = await self.build_allowance_report(query_params)
        elif report_type == "attendance_special_allowance":
            data = await self.build_attendance_special_allowance_report(query_params)
        elif report_type == "other_special_allowance":
            data = await self.build_other_special_allowance_report(query_params)
        elif report_type == "leave":
            data = await self.build_leave_report(query_params)
        elif report_type == "all":
            workbooks = await self.build_all_report(query_params)

            wb = openpyxl.Workbook()

            wb.remove(wb.active)

            for workbook in workbooks:
                for sheet in workbook.worksheets:
                    print(sheet.title)
                    new_sheet = wb.create_sheet(sheet.title)

                    for row in sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                            await self.copy_cell(cell, new_cell)

            data = wb

        else:
            raise HTTPException(status_code=501, detail="Not Implemented Yet!")

        output = io.BytesIO()

        data.save(output)

        output.seek(0)

        return output

    async def download_bank_salary_report(self, query_params):
        check_query_params = all(key in query_params for key in ["batch_id"])

        if not check_query_params:
            raise HTTPException(status_code=400, detail="Invalid query parameters")

        batch_id = query_params["batch_id"]

        obj = AdminCrud(self.payload, self.mongo_client)

        batch = await obj.get_bank_salary_batch(batch_id)

        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        employee_ids = batch["employee_ids"]
        # month = first_day_of_current_month()
        month = first_day_of_last_month()

        data = await self.build_bank_salary_report(employee_ids, month)
        return data
        raise HTTPException(status_code=501, detail="Not Implemented Yet!")

    async def build_bank_salary_report(self, employee_ids, month):
        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_bank_salary_data(employee_ids, month)

        wb = openpyxl.Workbook()
        ws = wb.active

        bold_font = Font(bold=True)

        rows = []
        ws.append(["IFSC", "ACCOUNT NO", "NAME", "NET SALARY", "MONTH OF SALARY"])

        total = 0
        for i in data:
            total += i["net_salary"]
            ws.append(
                [
                    i["bank_details"]["ifsc_code"],
                    i["bank_details"]["account_number"],
                    i["name"],
                    i["net_salary"],
                    month.strftime("%B - %Y"),
                ]
            )

        total_row_index = ws.max_row + 2
        ws.append(["", "", "", ""])
        ws.append(["", "", "Total", total, ""])
        for cell in ws[total_row_index]:
            cell.font = bold_font

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return output

    async def build_salary_report(self, query_params):
        try:
            datetime.datetime.strptime(query_params["month"], "%Y-%m")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Date format is incorrect. Expected format: YYYY-MM",
            )
        employee_id = query_params["employee_id"]

        emp = await employee_crud.get_employee(employee_id, self.mongo_client)

        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        month = first_day_of_current_month(
            month=int(query_params["month"].split("-")[1]),
            year=int(query_params["month"].split("-")[0]),
        )

        res, _ = await employee_crud.get_employee_with_computed_fields(
            query_params["employee_id"], self.mongo_client, month=month
        )

        data = {
            "Employee ID": res["basic_information"]["employee_id"],
            "Name": res["basic_information"]["name"],
            "Gross Salary": res["basic_salary"]["gross_salary"],
            "PF": res["basic_salary"]["pf"],
            "ESI": res["basic_salary"]["esi"],
            "Loan": res["loan_and_advance"]["loan"],
            "Loss of Pay": res["monthly_compensation"]["loss_of_pay"],
            "Salary Advance": res["loan_and_advance"]["salary_advance"],
            "Allowance": res["salary_incentives"]["allowance"],
            "Increment": res["salary_incentives"]["increment"],
            "Bonus": res["salary_incentives"]["bonus"],
            "Leave Cashback": res["monthly_compensation"]["leave_cashback"],
            "Last Year Leave Cashback": res["monthly_compensation"][
                "last_year_leave_cashback"
            ],
            "Attendance Special Allowance": res["monthly_compensation"][
                "attendance_special_allowance"
            ],
            "Other Special Allowance": res["monthly_compensation"][
                "other_special_allowance"
            ],
            "OT": res["monthly_compensation"]["overtime"],
            "Total Leave Days": res["leaves_and_permissions"]["total_leave_days"],
            "Monthly Leave Days": res["leaves_and_permissions"]["monthly_leave_days"],
            "Net Salary": res["basic_salary"]["net_salary"],
        }

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Salary Report"

        ws.append(
            [
                "Salary Report",
                " ",
                query_params["employee_id"],
                " ",
                datetime.datetime.strptime(query_params["month"], "%Y-%m").strftime(
                    "%B - %Y"
                ),
            ]
        )

        cell = ws.cell(row=1, column=1)
        cell.font = Font(bold=True)

        ws.append(["", ""])

        # Define the color fills
        green_font = Font(color="00FF00")
        red_font = Font(color="FF0000")

        # Fields that add to and subtract from the net salary
        add_fields = [
            "Gross Salary",
            "Allowance",
            "Increment",
            "Bonus",
            "Leave Cashback",
            "Last Year Leave Cashback",
            "Attendance Special Allowance",
            "Other Special Allowance",
            "OT",
        ]
        subtract_fields = ["PF", "ESI", "Loan", "Loss of Pay", "Salary Advance"]

        for key, value in data.items():
            ws.append([key, "", "", value])
            # Apply green font for add fields and red font for subtract fields
            # if key in add_fields:
            #     ws[ws.max_row][3].font = green_font  # Apply font color to the value cell
            # elif key in subtract_fields:
            #     ws[ws.max_row][3].font = red_font  # Apply font color to the value cell

            ws.append(["", ""])

        return wb

    async def build_increment_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_increment_report_data(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Increment Report"

        bold_font = Font(bold=True)

        ws.append(
            [
                "Employee ID",
                "Name",
                "Increment Amount",
                "Increment Month",
                "Increment Date",
                "Posted By",
            ]
        )

        total = 0
        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["amount"],
                    i["month"],
                    i["posted_at"],
                    i["posted_by"],
                ]
            )
            total += i["amount"]

        total_row_index = ws.max_row + 2

        ws.append(["", "", "", ""])
        ws.append(["", "Total", total, ""])

        for cell in ws[total_row_index]:
            cell.font = bold_font

        return wb

    async def build_bonus_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_bonus_report_data(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Bonus Report"

        bold_font = Font(bold=True)

        ws.append(
            [
                "Employee ID",
                "Name",
                "Bonus Amount",
                "Bonus Month",
                "Bonus Date",
                "Posted By",
            ]
        )

        total = 0
        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["amount"],
                    i["month"],
                    i["posted_at"],
                    i["posted_by"],
                ]
            )
            total += i["amount"]

        total_row_index = ws.max_row + 2

        ws.append(["", "", "", ""])
        ws.append(["", "Total", total, ""])

        for cell in ws[total_row_index]:
            cell.font = bold_font

        return wb

    async def build_allowance_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_allowance_report_data(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Allowance Report"

        bold_font = Font(bold=True)

        ws.append(
            [
                "Employee ID",
                "Name",
                "Allowance Amount",
                "Allowance Month",
                "Allowance Date",
                "Posted By",
            ]
        )

        total = 0
        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["amount"],
                    i["month"],
                    i["posted_at"],
                    i["posted_by"],
                ]
            )
            total += i["amount"]

        total_row_index = ws.max_row + 2

        ws.append(["", "", "", ""])
        ws.append(["", "Total", total, ""])

        for cell in ws[total_row_index]:
            cell.font = bold_font

        return wb

    async def build_attendance_special_allowance_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_attendance_special_allowance_report_data(
            start_date, end_date
        )

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Attendance Special Allowance Report"

        bold_font = Font(bold=True)

        ws.append(
            [
                "Employee ID",
                "Name",
                "Allowance Amount",
                "Allowance Month",
                "Allowance Date",
                "Posted By",
            ]
        )

        total = 0

        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["amount"],
                    i["month"],
                    i["posted_at"],
                    i["posted_by"],
                ]
            )
            total += i["amount"]

        total_row_index = ws.max_row + 2

        ws.append(["", "", "", ""])
        ws.append(["", "Total", total, ""])

        for cell in ws[total_row_index]:
            cell.font = bold_font

        return wb

    async def build_other_special_allowance_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_other_special_allowance_report_data(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Other Special Allowance Report"

        bold_font = Font(bold=True)

        ws.append(
            [
                "Employee ID",
                "Name",
                "Allowance Amount",
                "Allowance Month",
                "Allowance Date",
                "Posted By",
            ]
        )

        total = 0
        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["amount"],
                    i["month"],
                    i["posted_at"],
                    i["posted_by"],
                ]
            )
            total += i["amount"]

        total_row_index = ws.max_row + 2

        ws.append(["", "", "", ""])
        ws.append(["", "Total", total, ""])

        for cell in ws[total_row_index]:
            cell.font = bold_font

        return wb

    async def build_leave_report(self, query_params):
        start_year = int(query_params["start_year"])
        end_year = int(query_params["end_year"])

        if start_year > end_year:
            raise HTTPException(
                status_code=400, detail="Start year cannot be greater than end year"
            )

        start_date = datetime.datetime(start_year, 4, 1, 0, 0, 0)
        end_date = datetime.datetime(end_year, 3, 31, 0, 0, 0)

        obj = AdminCrud(self.payload, self.mongo_client)

        data = await obj.get_leave_report_data(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Leave Report"

        ws.append(
            [
                "Employee ID",
                "Name",
                "Leave Type",
                "Start Date",
                "End Date",
                "No of Days",
                "Leave Month",
                "Approved By",
                "Reason",
                "Remarks",
            ]
        )

        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["leave_type"].capitalize(),
                    i["start_date"],
                    i["end_date"],
                    i["no_of_days"],
                    i["month"],
                    i["approved_by"],
                    i["reason"],
                    i["remarks"],
                ]
            )

        wb.create_sheet("Permission Report")
        # Switch to the second sheet
        ws = wb["Permission Report"]

        data = await obj.get_permission_report_data(start_date, end_date)

        ws.title = "Permission Report"

        ws.append(
            [
                "Employee ID",
                "Name",
                "Date",
                "Start Time",
                "End Time",
                "No of Hours",
                "Permission Month",
                "Approved By",
                "Reason",
                "Remarks",
            ]
        )

        for i in data:
            ws.append(
                [
                    i["employee_id"],
                    i["name"],
                    i["date"],
                    i["start_time"],
                    i["end_time"],
                    i["no_of_hours"],
                    i["month"],
                    i["approved_by"],
                    i["reason"],
                    i["remarks"],
                ]
            )

        return wb

    async def build_all_report(self, query_params):
        wb = []
        print(query_params)
        # wb.append(await self.build_salary_report(query_params))
        wb.append(await self.build_increment_report(query_params))
        wb.append(await self.build_bonus_report(query_params))
        wb.append(await self.build_allowance_report(query_params))
        wb.append(await self.build_attendance_special_allowance_report(query_params))
        wb.append(await self.build_other_special_allowance_report(query_params))
        wb.append(await self.build_leave_report(query_params))
        print(wb)
        return wb
